import os
from datetime import datetime

from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

PLANILHA = os.path.join(
    BASE_DIR,
    "Planilha_Mestra.xlsx"
)


def normalizar_cabecalho(valor):
    if valor is None:
        return ""

    return str(valor).strip().lower()


def registrar_aprovado(dados, observacao="Cadastro aprovado"):

    if not os.path.exists(PLANILHA):
        raise FileNotFoundError(
            f"Planilha não encontrada: {PLANILHA}"
        )

    workbook = load_workbook(PLANILHA)

    print(
        "Abas encontradas:",
        workbook.sheetnames
    )

    # Usa a primeira aba da planilha.
    sheet = workbook.active

    print(
        f"Preenchendo a aba: {sheet.title}"
    )

    # Lê os cabeçalhos da primeira linha
    cabecalhos = {}

    for coluna in range(1, sheet.max_column + 1):

        valor = sheet.cell(
            row=1,
            column=coluna
        ).value

        if valor:
            cabecalhos[
                normalizar_cabecalho(valor)
            ] = coluna

    print(
        "Cabeçalhos encontrados:",
        cabecalhos
    )

    # Mapeamento entre possíveis nomes da planilha
    mapeamento = {

        "cpf": [
            "cpf"
        ],

        "nome": [
            "nome",
            "nome completo",
            "cliente"
        ],

        "data_nascimento": [
            "data de nascimento",
            "data nascimento",
            "nascimento"
        ],

        "endereco": [
            "endereço",
            "endereco"
        ],

        "email": [
            "e-mail",
            "email"
        ],

        "telefone": [
            "telefone",
            "celular",
            "telefone/celular"
        ],

        "status": [
            "status",
            "situação",
            "situacao"
        ],

        "data_processamento": [
            "data de processamento",
            "data processamento",
            "data cadastro"
        ],

        "observacao": [
            "observação",
            "observacao",
            "observações",
            "observacoes"
        ]
    }

    # Próxima linha vazia
    # Encontra a primeira linha realmente vazia
    linha = 2

    while True:
        cpf_coluna = cabecalhos.get("cpf")
        nome_coluna = cabecalhos.get("nome")

        cpf_existente = (
            sheet.cell(linha, cpf_coluna).value
            if cpf_coluna
            else None
        )

        nome_existente = (
            sheet.cell(linha, nome_coluna).value
            if nome_coluna
            else None
        )

        if not cpf_existente and not nome_existente:
            break

        linha += 1

    # Se a planilha tiver apenas a linha de cabeçalho,
    # max_row será 1 e a gravação ocorrerá na linha 2.

    for campo, possibilidades in mapeamento.items():

        valor = dados.get(campo)

        if campo == "status":
            valor = "APROVADO"

        elif campo == "data_processamento":
            valor = datetime.now().strftime(
                "%d/%m/%Y %H:%M:%S"
            )

        elif campo == "observacao":
            valor = observacao

        if valor is None:
            continue

        for nome_coluna in possibilidades:

            nome_coluna = normalizar_cabecalho(
                nome_coluna
            )

            if nome_coluna in cabecalhos:

                coluna = cabecalhos[
                    nome_coluna
                ]

                sheet.cell(
                    row=linha,
                    column=coluna
                ).value = valor

                break

    workbook.save(PLANILHA)

    print(
        f"Registro salvo na Planilha Mestra "
        f"(linha {linha})."
    )